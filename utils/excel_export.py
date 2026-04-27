# -*- coding: utf-8 -*-
"""
Экспорт отчётов в Excel.
"""

import io
from typing import Dict, Any
from datetime import datetime

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


class ExcelReportExporter:
    """Экспорт отчётов в формат Excel."""

    def __init__(self):
        if not EXCEL_AVAILABLE:
            raise ImportError("Требуется установить pandas и openpyxl: pip install pandas openpyxl")

    def export_daily_report(self, report: Dict[str, Any]) -> bytes:
        """Экспорт дневного отчёта в Excel."""
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Основной лист со статистикой
            summary_data = {
                'Параметр': ['Всего проверок', 'Годен (OK)', 'Брак (NG)', '% качества'],
                'Значение': [
                    report.get('total', 0),
                    report.get('ok_count', 0),
                    report.get('ng_count', 0),
                    f"{report.get('ok_percent', 0)}%"
                ]
            }
            df_summary = pd.DataFrame(summary_data)
            df_summary.to_excel(writer, sheet_name='Сводка', index=False)
            
            # Лист с почасовой статистикой
            if report.get('hourly_stats'):
                df_hourly = pd.DataFrame(report['hourly_stats'])
                df_hourly = df_hourly.rename(columns={
                    'hour': 'Час',
                    'total': 'Всего',
                    'ok': 'OK',
                    'ng': 'NG',
                    'ok_percent': '% качества'
                })
                df_hourly.to_excel(writer, sheet_name='По часам', index=False)
            
            # Лист с проектами
            if report.get('project_stats'):
                df_projects = pd.DataFrame(report['project_stats'])
                df_projects = df_projects.rename(columns={
                    'project': 'Проект',
                    'total': 'Всего',
                    'ok': 'OK',
                    'ng': 'NG',
                    'ok_percent': '% качества'
                })
                df_projects.to_excel(writer, sheet_name='Проекты', index=False)
            
            # Лист с результатами
            if report.get('results'):
                df_results = pd.DataFrame(report['results'])
                if 'timestamp' in df_results.columns:
                    df_results['timestamp'] = pd.to_datetime(df_results['timestamp'])
                df_results = df_results.rename(columns={
                    'timestamp': 'Время',
                    'result': 'Результат',
                    'project_name': 'Проект',
                    'scenario': 'Сценарий'
                })
                cols_to_keep = ['Время', 'Результат', 'Проект', 'Сценарий']
                df_results = df_results[[c for c in cols_to_keep if c in df_results.columns]]
                df_results.to_excel(writer, sheet_name='Результаты', index=False)
        
        # Применяем стили
        output.seek(0)
        return self._apply_styles(output, f"Дневной отчёт за {report.get('date', '')}")

    def export_shift_report(self, report: Dict[str, Any]) -> bytes:
        """Экспорт сменного отчёта в Excel."""
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            summary_data = {
                'Параметр': ['Дата', 'Смена', 'Всего проверок', 'Годен (OK)', 'Брак (NG)', '% качества'],
                'Значение': [
                    report.get('date', ''),
                    report.get('shift_name', ''),
                    report.get('total', 0),
                    report.get('ok_count', 0),
                    report.get('ng_count', 0),
                    f"{report.get('ok_percent', 0)}%"
                ]
            }
            df_summary = pd.DataFrame(summary_data)
            df_summary.to_excel(writer, sheet_name='Сводка', index=False)
            
            if report.get('results'):
                df_results = pd.DataFrame(report['results'])
                if 'timestamp' in df_results.columns:
                    df_results['timestamp'] = pd.to_datetime(df_results['timestamp'])
                df_results = df_results.rename(columns={
                    'timestamp': 'Время',
                    'result': 'Результат'
                })
                df_results.to_excel(writer, sheet_name='Результаты', index=False)
        
        output.seek(0)
        return self._apply_styles(output, f"Сменный отчёт за {report.get('date', '')}")

    def export_weekly_report(self, report: Dict[str, Any]) -> bytes:
        """Экспорт недельного отчёта в Excel."""
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            summary_data = {
                'Параметр': ['Неделя', 'Период', 'Всего проверок', 'Годен (OK)', 'Брак (NG)', '% качества'],
                'Значение': [
                    report.get('week', ''),
                    f"{report.get('date_from', '')} — {report.get('date_to', '')}",
                    report.get('total', 0),
                    report.get('ok_count', 0),
                    report.get('ng_count', 0),
                    f"{report.get('ok_percent', 0)}%"
                ]
            }
            df_summary = pd.DataFrame(summary_data)
            df_summary.to_excel(writer, sheet_name='Сводка', index=False)
            
            if report.get('daily_stats'):
                df_daily = pd.DataFrame(report['daily_stats'])
                df_daily = df_daily.rename(columns={
                    'date': 'Дата',
                    'total': 'Всего',
                    'ok': 'OK',
                    'ng': 'NG',
                    'ok_percent': '% качества'
                })
                df_daily.to_excel(writer, sheet_name='По дням', index=False)
        
        output.seek(0)
        return self._apply_styles(output, f"Недельный отчёт #{report.get('week', '')}")

    def export_monthly_report(self, report: Dict[str, Any]) -> bytes:
        """Экспорт месячного отчёта в Excel."""
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            summary_data = {
                'Параметр': ['Месяц', 'Год', 'Всего проверок', 'Годен (OK)', 'Брак (NG)', '% качества'],
                'Значение': [
                    report.get('month_name', ''),
                    report.get('year', ''),
                    report.get('total', 0),
                    report.get('ok_count', 0),
                    report.get('ng_count', 0),
                    f"{report.get('ok_percent', 0)}%"
                ]
            }
            df_summary = pd.DataFrame(summary_data)
            df_summary.to_excel(writer, sheet_name='Сводка', index=False)
            
            if report.get('daily_stats'):
                df_daily = pd.DataFrame(report['daily_stats'])
                df_daily = df_daily.rename(columns={
                    'date': 'Дата',
                    'total': 'Всего',
                    'ok': 'OK',
                    'ng': 'NG',
                    'ok_percent': '% качества'
                })
                df_daily.to_excel(writer, sheet_name='По дням', index=False)
            
            if report.get('shift_stats'):
                df_shifts = pd.DataFrame(report['shift_stats'])
                df_shifts = df_shifts.rename(columns={
                    'shift_name': 'Смена',
                    'total': 'Всего',
                    'ok': 'OK',
                    'ng': 'NG',
                    'ok_percent': '% качества'
                })
                df_shifts.to_excel(writer, sheet_name='По сменам', index=False)
        
        output.seek(0)
        return self._apply_styles(output, f"Месячный отчёт {report.get('month_name', '')} {report.get('year', '')}")

    def export_ng_analysis(self, report: Dict[str, Any]) -> bytes:
        """Экспорт анализа брака в Excel."""
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            summary_data = {
                'Параметр': ['Период', 'Всего брака'],
                'Значение': [
                    f"{report.get('date_from', '')} — {report.get('date_to', '')}",
                    report.get('total_ng', 0)
                ]
            }
            df_summary = pd.DataFrame(summary_data)
            df_summary.to_excel(writer, sheet_name='Сводка', index=False)
            
            if report.get('time_distribution'):
                time_data = {
                    'Время суток': ['Утро (06:00-14:00)', 'День (14:00-22:00)', 'Ночь (22:00-06:00)'],
                    'Количество брака': [
                        report['time_distribution'].get('morning', 0),
                        report['time_distribution'].get('afternoon', 0),
                        report['time_distribution'].get('night', 0)
                    ]
                }
                df_time = pd.DataFrame(time_data)
                df_time.to_excel(writer, sheet_name='По времени', index=False)
            
            if report.get('project_distribution'):
                df_projects = pd.DataFrame(report['project_distribution'])
                df_projects = df_projects.rename(columns={
                    'project': 'Проект',
                    'ng': 'Брак'
                })
                df_projects.to_excel(writer, sheet_name='По проектам', index=False)
        
        output.seek(0)
        return self._apply_styles(output, f"Анализ брака {report.get('date_from', '')}")

    def _apply_styles(self, output: io.BytesIO, title: str) -> bytes:
        """Применение стилей к Excel файлу."""
        from openpyxl import load_workbook
        
        wb = load_workbook(output)
        
        # Стили
        title_font = Font(bold=True, size=14, color="FFFFFF")
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for ws in wb.worksheets:
            # Заголовок
            ws.insert_rows(1)
            max_col = ws.max_column
            ws.merge_cells(f'A1:{chr(64 + max_col)}1' if max_col <= 26 else f'A1:Z1')
            cell = ws['A1']
            cell.value = title
            cell.font = title_font
            cell.fill = PatternFill(start_color="004080", end_color="004080", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Стили заголовков таблицы (строка 2 после вставки)
            for col in range(1, max_col + 1):
                try:
                    header_cell = ws.cell(row=2, column=col)
                    # Пропускаем объединённые ячейки
                    if not isinstance(header_cell, type(ws.merged_cells.ranges[0].cells[0]) if ws.merged_cells.ranges else object):
                        header_cell.font = header_font
                        header_cell.fill = header_fill
                        header_cell.alignment = header_alignment
                        header_cell.border = thin_border
                except:
                    pass
            
            # Стили ячеек данных
            for row in range(3, ws.max_row + 1):
                for col in range(1, max_col + 1):
                    try:
                        cell = ws.cell(row=row, column=col)
                        cell.border = thin_border
                    except:
                        pass
            
            # Автоширина колонок - исправленная версия
            for col_idx in range(1, max_col + 1):
                max_length = 0
                column_letter = ws.cell(row=2, column=col_idx).column_letter
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Сохраняем
        output_new = io.BytesIO()
        wb.save(output_new)
        output_new.seek(0)
        return output_new.read()


# Глобальный экземпляр
_exporter: ExcelReportExporter = None


def get_excel_exporter() -> ExcelReportExporter:
    """Получение экземпляра ExcelReportExporter."""
    global _exporter
    if _exporter is None:
        _exporter = ExcelReportExporter()
    return _exporter
